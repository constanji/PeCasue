"""Agent toolbox — shared by Copilot, Mapper, Password, UnknownChannel agents.

Each tool is a pure function that takes a JSON-serialisable payload and
returns a JSON-serialisable result. They never touch the LLM directly; that's
the agent's job. Heavy operations are intentionally bounded (head=20 rows etc).
"""
from __future__ import annotations

import csv
import json
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from server.core.paths import (
    get_agent_drafts_dir,
    get_channel_run_dir,
    get_task_dir,
    get_task_extracted_dir,
    get_task_log_path,
    get_channel_log_path,
)
from server.core.pipeline_state import (
    AgentInteraction,
    StateManager,
    TaskState,
)
from server.core.task_logger import task_log
from server.core.task_repo import TaskRepo
from server.rules import store as rule_store
from server.rules.password_book import lookup_password as pb_lookup
from server.rules.schema import RuleKind


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------- file / log readers ----------


def _resolve_file_path(
    task_id: str,
    rel_path: str,
    channel_id: Optional[str] = None,
) -> Path:
    """Resolve a rel_path to an absolute path, trying multiple roots for robustness.

    Resolution order:
      1. If ``rel_path`` already starts with ``extracted/``, always resolve from task root.
      2. If ``channel_id`` given: try ``extracted/{channel_id}/rel_path`` first.
      3. Try task root (``tasks/{task_id}/rel_path``) — covers paths from list_task_files without channel_id.
      4. If channel_id given and still not found, try searching all extracted channels
         for a file whose suffix matches ``rel_path``.
    """
    task_root = get_task_dir(task_id).resolve()

    # Case 1: rel_path already contains the full extracted/... path from task root
    if rel_path.startswith("extracted/") or rel_path.startswith("channels/"):
        candidate = (task_root / rel_path).resolve()
        if str(candidate).startswith(str(task_root)) and candidate.exists():
            return candidate

    # Case 2: channel_id given — try extracted/{channel_id}/ first
    if channel_id:
        ch_root = get_task_extracted_dir(task_id, channel_id).resolve()
        candidate = (ch_root / rel_path).resolve()
        if str(candidate).startswith(str(task_root)) and candidate.exists():
            return candidate

    # Case 3: try task root
    candidate = (task_root / rel_path).resolve()
    if str(candidate).startswith(str(task_root)) and candidate.exists():
        return candidate

    # Case 4: try stripping/adding channel prefix
    # If rel_path looks like "extracted/{some_channel}/rest", try resolving from task root
    parts = rel_path.replace("\\", "/").split("/")
    if len(parts) >= 2 and parts[0] == "extracted":
        candidate = (task_root / rel_path).resolve()
        if str(candidate).startswith(str(task_root)) and candidate.exists():
            return candidate

    # Case 5: search across all extracted channels for matching suffix
    if channel_id:
        extracted_root = get_task_extracted_dir(task_id, None).resolve()
        suffix = Path(rel_path).name
        for ch_dir in extracted_root.iterdir():
            if not ch_dir.is_dir():
                continue
            for found in ch_dir.rglob(suffix):
                try:
                    rel = found.relative_to(ch_dir)
                    if str(rel) == rel_path:
                        return found.resolve()
                except ValueError:
                    continue

    # Return the best guess (will fail with clear error if not found)
    if channel_id:
        return (get_task_extracted_dir(task_id, channel_id) / rel_path).resolve()
    return (task_root / rel_path).resolve()


def list_task_files(task_id: str, channel_id: Optional[str] = None) -> Dict[str, Any]:
    """List files under the task folder.

    **All ``rel_path`` values are relative to the TASK ROOT** (``tasks/{task_id}/``)
    regardless of whether ``channel_id`` is passed. This ensures the paths can be
    used directly by ``read_csv``, ``read_excel``, etc. without needing to manually
    prepend or strip the channel prefix.
    """
    task_root = get_task_dir(task_id)
    if not task_root.exists():
        return {"task_id": task_id, "channel_id": channel_id, "files": []}

    if channel_id:
        scan_root = get_task_extracted_dir(task_id, channel_id)
    else:
        scan_root = task_root

    if not scan_root.exists():
        return {"task_id": task_id, "channel_id": channel_id, "files": []}

    files = []
    for p in sorted(scan_root.rglob("*")):
        if p.is_file() and not p.name.startswith("."):
            try:
                # Always return path relative to task root, not scan_root
                rel = str(p.resolve().relative_to(task_root.resolve()))
            except ValueError:
                rel = str(p)
            files.append({"rel_path": rel, "size": p.stat().st_size})
    return {
        "task_id": task_id,
        "channel_id": channel_id,
        "files": files,
        "note": "All rel_path values are relative to the task root. Use them directly with read_csv/read_excel (no need to add channel prefix).",
    }


def read_text(task_id: str, rel_path: str, *, channel_id: Optional[str] = None, head: int = 50) -> Dict[str, Any]:
    """Read text file. ``rel_path`` is relative to the task root (consistent with ``list_task_files``)."""
    p = _resolve_file_path(task_id, rel_path, channel_id)
    task_root = get_task_dir(task_id).resolve()
    if not str(p).startswith(str(task_root)):
        return {"error": "path escapes task scope"}
    if not p.exists() or not p.is_file():
        return {"error": f"file not found: {rel_path}"}
    try:
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            lines = []
            for i, line in enumerate(f):
                if i >= head:
                    break
                lines.append(line.rstrip("\n"))
        return {"rel_path": rel_path, "lines": lines, "head": head}
    except Exception as exc:  # noqa: BLE001
        return {"error": str(exc)}


def read_csv(task_id: str, rel_path: str, *, channel_id: Optional[str] = None, head: int = 20) -> Dict[str, Any]:
    """Read CSV header + rows. ``rel_path`` is relative to the task root (consistent with ``list_task_files``)."""
    p = _resolve_file_path(task_id, rel_path, channel_id)
    task_root = get_task_dir(task_id).resolve()
    if not str(p).startswith(str(task_root)):
        return {"error": "path escapes task scope"}
    if not p.exists():
        return {"error": f"file not found: {rel_path}"}
    out_rows: List[Dict[str, Any]] = []
    columns: List[str] = []
    try:
        with open(p, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            r = csv.reader(f)
            for i, row in enumerate(r):
                if i == 0:
                    columns = row
                    continue
                if i > head:
                    break
                out_rows.append({c: (row[j] if j < len(row) else "") for j, c in enumerate(columns)})
    except Exception as exc:  # noqa: BLE001
        return {"error": str(exc)}
    return {"rel_path": rel_path, "columns": columns, "rows": out_rows, "head": head}


def read_excel(
    task_id: str,
    rel_path: str,
    *,
    channel_id: Optional[str] = None,
    sheet: Optional[str] = None,
    head: int = 20,
) -> Dict[str, Any]:
    """Read Excel header + rows. ``rel_path`` is relative to the task root (consistent with ``list_task_files``)."""
    if rel_path.lower().endswith((".csv", ".tsv", ".txt", ".log", ".json")):
        return read_csv(task_id, rel_path, channel_id=channel_id, head=head)
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl is not installed; cannot read xlsx"}
    p = _resolve_file_path(task_id, rel_path, channel_id)
    task_root = get_task_dir(task_id).resolve()
    if not str(p).startswith(str(task_root)):
        return {"error": "path escapes task scope"}
    if not p.exists():
        return {"error": f"file not found: {rel_path}"}
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows: List[Dict[str, Any]] = []
        columns: List[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                columns = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                continue
            if i > head:
                break
            rows.append(
                {c: (row[j] if j < len(row) else None) for j, c in enumerate(columns)}
            )
        wb.close()
        return {
            "rel_path": rel_path,
            "sheet": sheet or (ws.title if ws else None),
            "columns": columns,
            "rows": rows,
            "head": head,
        }
    except Exception as exc:  # noqa: BLE001
        return {"error": str(exc)}


def read_verify_summary(task_id: str, channel_id: str, run_id: Optional[str] = None) -> Dict[str, Any]:
    state = StateManager.load_state(task_id)
    if not state or channel_id not in state.channels:
        return {"error": "task or channel not found"}
    ch = state.channels[channel_id]
    if not ch.runs:
        return {"runs": [], "summary": None}
    if run_id is None:
        run = ch.runs[-1]
    else:
        run = next((r for r in ch.runs if r.run_id == run_id), None)
        if run is None:
            return {"error": "run not found"}
    return {"run_id": run.run_id, "status": str(run.status), "summary": run.verify_summary}


def read_log(task_id: str, *, channel_id: Optional[str] = None, tail: int = 200) -> Dict[str, Any]:
    p = get_channel_log_path(task_id, channel_id) if channel_id else get_task_log_path(task_id)
    if not p.exists():
        return {"lines": []}
    with open(p, "r", encoding="utf-8", errors="replace") as f:
        all_lines = f.readlines()
    return {"lines": [ln.rstrip("\n") for ln in all_lines[-tail:]]}


# ---------- rule queries ----------


def query_rules(kind: str, *, filter: Optional[Dict[str, Any]] = None, limit: int = 50) -> Dict[str, Any]:
    try:
        rk = RuleKind(kind)
    except ValueError:
        return {"error": f"unknown rule kind: {kind}"}
    if rk == RuleKind.PASSWORD_BOOK:
        return {"error": "password_book is not queryable via this tool"}
    table = rule_store.load_rule(rk)
    rows = table.rows
    if filter:
        rows = [
            r
            for r in rows
            if all(str(r.get(k)) == str(v) for k, v in filter.items())
        ]
    return {
        "kind": rk.value,
        "columns": table.columns,
        "rows": rows[:limit],
        "matched": len(rows),
    }


def lookup_password(*, scope: Optional[str] = None, pattern: Optional[str] = None) -> Dict[str, Any]:
    pw = pb_lookup(scope=scope, pattern=pattern)
    if not pw:
        return {"found": False}
    return {"found": True, "password": pw}


# ---------- agent drafts (write proposals — Human must approve) ----------


def _draft_dir(task_id: str) -> Path:
    d = get_agent_drafts_dir(task_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def list_drafts(task_id: str) -> Dict[str, Any]:
    d = _draft_dir(task_id)
    out = []
    for p in sorted(d.glob("*.json")):
        try:
            payload = json.loads(p.read_text(encoding="utf-8"))
            payload["_id"] = p.stem
            out.append(payload)
        except Exception:
            continue
    return {"task_id": task_id, "drafts": out}


def _persist_draft(task_id: str, payload: Dict[str, Any]) -> str:
    d = _draft_dir(task_id)
    draft_id = uuid.uuid4().hex[:12]
    payload = {
        **payload,
        "draft_id": draft_id,
        "created_at": _utc_now(),
        "status": "pending",
    }
    (d / f"{draft_id}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return draft_id


def propose_rule_patch(
    task_id: str,
    *,
    kind: str,
    patch: Dict[str, Any],
    rationale: Optional[str] = None,
) -> Dict[str, Any]:
    draft_id = _persist_draft(
        task_id,
        {
            "kind": "rule_patch",
            "rule_kind": kind,
            "patch": patch,
            "rationale": rationale,
        },
    )
    TaskRepo.append_event(
        task_id,
        "agent.draft.proposed",
        payload={"draft_id": draft_id, "kind": "rule_patch", "rule_kind": kind},
    )
    return {"draft_id": draft_id, "status": "pending"}


def propose_replace_file(
    task_id: str,
    *,
    channel_id: str,
    rel_path: str,
    new_content_uri: str,
    rationale: Optional[str] = None,
) -> Dict[str, Any]:
    draft_id = _persist_draft(
        task_id,
        {
            "kind": "replace_file",
            "channel_id": channel_id,
            "rel_path": rel_path,
            "new_content_uri": new_content_uri,
            "rationale": rationale,
        },
    )
    TaskRepo.append_event(
        task_id,
        "agent.draft.proposed",
        payload={"draft_id": draft_id, "kind": "replace_file", "channel_id": channel_id},
    )
    return {"draft_id": draft_id, "status": "pending"}


def mark_row_resolved(
    task_id: str,
    *,
    channel_id: str,
    run_id: str,
    verify_row_id: str,
    reason: str,
    actor: str = "agent",
) -> Dict[str, Any]:
    state = StateManager.load_state(task_id)
    if not state or channel_id not in state.channels:
        return {"error": "task or channel not found"}
    ch = state.channels[channel_id]
    run = next((r for r in ch.runs if r.run_id == run_id), None)
    if run is None or not run.verify_summary:
        return {"error": "run or verify summary not found"}
    rows = run.verify_summary.get("rows", [])
    found = False
    for row in rows:
        if row.get("row_id") == verify_row_id:
            row["resolved"] = True
            row["resolution_reason"] = reason
            row["resolved_at"] = _utc_now()
            found = True
            break
    if not found:
        return {"error": "verify row not found"}
    run.agent_interactions.append(
        AgentInteraction(
            kind="resolve",
            summary=f"resolved {verify_row_id}: {reason}",
            payload={"verify_row_id": verify_row_id, "reason": reason, "actor": actor},
        )
    )
    StateManager.save_state(state)
    task_log(task_id, f"Agent resolved {verify_row_id}: {reason}", channel=channel_id)
    return {"resolved": True, "verify_row_id": verify_row_id}


# ---------- data analysis tools ----------


def _load_tabular_rows(
    task_id: str,
    rel_path: str,
    channel_id: Optional[str] = None,
    sheet: Optional[str] = None,
) -> Dict[str, Any]:
    """Internal helper: load all rows from a CSV/Excel file as list[dict]."""
    p = _resolve_file_path(task_id, rel_path, channel_id)
    task_root = get_task_dir(task_id).resolve()
    if not str(p).startswith(str(task_root)):
        return {"error": "path escapes task scope"}
    if not p.exists():
        return {"error": f"file not found: {rel_path}"}

    rows: List[Dict[str, Any]] = []
    columns: List[str] = []
    try:
        if rel_path.lower().endswith((".csv", ".tsv", ".txt")):
            with open(p, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
                r = csv.reader(f)
                for i, row in enumerate(r):
                    if i == 0:
                        columns = row
                        continue
                    rows.append(
                        {c: (row[j] if j < len(row) else "") for j, c in enumerate(columns)}
                    )
        else:
            from openpyxl import load_workbook

            wb = load_workbook(p, data_only=True, read_only=True)
            ws = wb[sheet] if sheet else wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    columns = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                    continue
                rows.append(
                    {c: (row[j] if j < len(row) else None) for j, c in enumerate(columns)}
                )
            wb.close()
    except Exception as exc:  # noqa: BLE001
        return {"error": str(exc)}
    return {"columns": columns, "rows": rows, "total": len(rows)}


def _try_numeric(val: Any) -> Any:
    """Attempt to convert a value to a number for aggregation."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def filter_table(
    task_id: str,
    rel_path: str,
    *,
    column: str,
    op: str,
    value: str,
    channel_id: Optional[str] = None,
    sheet: Optional[str] = None,
    head: int = 50,
) -> Dict[str, Any]:
    """Filter CSV/Excel rows. Parameter names MUST match Python exactly:

    ``task_id``, ``rel_path`` (not path/file_path), optional ``channel_id``, ``sheet``,
    keyword-only ``column`` (not column_name/search_column), ``op`` (not operator; values:
    eq, ne, contains, gt, lt, gte, lte, startswith, endswith), ``value`` (string),
    optional ``head``.

    Returns up to ``head`` matching rows.
    """
    data = _load_tabular_rows(task_id, rel_path, channel_id=channel_id, sheet=sheet)
    if "error" in data:
        return data
    columns = data["columns"]
    rows = data["rows"]

    if column not in columns:
        return {"error": f"column '{column}' not found; available: {columns[:20]}"}

    matches: List[Dict[str, Any]] = []
    for row in rows:
        cell = row.get(column)
        cell_s = str(cell) if cell is not None else ""
        val_s = str(value)

        ok = False
        if op == "eq":
            ok = cell_s == val_s
        elif op == "ne":
            ok = cell_s != val_s
        elif op == "contains":
            ok = val_s.lower() in cell_s.lower()
        elif op == "startswith":
            ok = cell_s.lower().startswith(val_s.lower())
        elif op == "endswith":
            ok = cell_s.lower().endswith(val_s.lower())
        elif op in ("gt", "lt", "gte", "lte"):
            cn = _try_numeric(cell)
            vn = _try_numeric(value)
            if cn is not None and vn is not None:
                if op == "gt":
                    ok = cn > vn
                elif op == "lt":
                    ok = cn < vn
                elif op == "gte":
                    ok = cn >= vn
                elif op == "lte":
                    ok = cn <= vn
        else:
            return {"error": f"unsupported operator: {op}"}

        if ok:
            matches.append(row)
            if len(matches) >= head:
                break

    return {
        "rel_path": rel_path,
        "column": column,
        "op": op,
        "value": value,
        "total_rows": len(rows),
        "matched": len(matches),
        "rows": matches,
        "head": head,
    }


def aggregate_table(
    task_id: str,
    rel_path: str,
    *,
    agg_fn: str,
    column: Optional[str] = None,
    group_by: Optional[str] = None,
    channel_id: Optional[str] = None,
    sheet: Optional[str] = None,
) -> Dict[str, Any]:
    """Aggregate CSV/Excel. Keywords: ``task_id``, ``rel_path``, ``agg_fn``; optional ``column``, ``group_by``, ``channel_id``, ``sheet``.

    agg_fn ∈ sum, count, avg, min, max, count_distinct. Non-count aggs require ``column``.
    """
    data = _load_tabular_rows(task_id, rel_path, channel_id=channel_id, sheet=sheet)
    if "error" in data:
        return data
    columns = data["columns"]
    rows = data["rows"]

    if agg_fn not in ("sum", "count", "avg", "min", "max", "count_distinct"):
        return {"error": f"unsupported agg_fn: {agg_fn}"}

    def _do_agg(row_list: List[Dict[str, Any]]) -> Any:
        if agg_fn == "count":
            return len(row_list)
        if agg_fn == "count_distinct":
            return len({str(r.get(column)) for r in row_list})
        if column is None:
            return {"error": f"agg_fn '{agg_fn}' requires a column"}
        nums = [_try_numeric(r.get(column)) for r in row_list]
        nums = [n for n in nums if n is not None]
        if not nums:
            return None
        if agg_fn == "sum":
            return round(sum(nums), 4)
        if agg_fn == "avg":
            return round(sum(nums) / len(nums), 4)
        if agg_fn == "min":
            return min(nums)
        if agg_fn == "max":
            return max(nums)
        return None

    if group_by:
        if group_by not in columns:
            return {"error": f"group_by column '{group_by}' not found; available: {columns[:20]}"}
        groups: Dict[str, List[Dict[str, Any]]] = {}
        for r in rows:
            key = str(r.get(group_by, ""))
            groups.setdefault(key, []).append(r)
        result = {}
        for key, grp in sorted(groups.items()):
            result[key] = {"count": len(grp), "value": _do_agg(grp)}
        return {
            "rel_path": rel_path,
            "agg_fn": agg_fn,
            "column": column,
            "group_by": group_by,
            "total_rows": len(rows),
            "groups": result,
        }
    else:
        return {
            "rel_path": rel_path,
            "agg_fn": agg_fn,
            "column": column,
            "total_rows": len(rows),
            "value": _do_agg(rows),
        }


def compare_files(
    task_id: str,
    path_a: str,
    path_b: str,
    *,
    channel_id: Optional[str] = None,
    key_column: Optional[str] = None,
    head: int = 30,
) -> Dict[str, Any]:
    """Compare two CSV/Excel files and report differences.

    Without ``key_column``: reports row-count and column-set differences.
    With ``key_column``: joins on that column and reports per-row value diffs.
    """
    data_a = _load_tabular_rows(task_id, path_a, channel_id=channel_id)
    if "error" in data_a:
        return {"error_a": data_a["error"]}
    data_b = _load_tabular_rows(task_id, path_b, channel_id=channel_id)
    if "error" in data_b:
        return {"error_b": data_b["error"]}

    cols_a = set(data_a["columns"])
    cols_b = set(data_b["columns"])
    summary: Dict[str, Any] = {
        "path_a": path_a,
        "path_b": path_b,
        "rows_a": len(data_a["rows"]),
        "rows_b": len(data_b["rows"]),
        "columns_only_in_a": sorted(cols_a - cols_b),
        "columns_only_in_b": sorted(cols_b - cols_a),
        "common_columns": sorted(cols_a & cols_b),
    }

    if not key_column:
        # Row-count level comparison only
        return summary

    # Build lookup on key_column
    def _key(row: Dict[str, Any]) -> str:
        return str(row.get(key_column, ""))

    lookup_a: Dict[str, Dict[str, Any]] = {}
    for r in data_a["rows"]:
        lookup_a[_key(r)] = r
    lookup_b: Dict[str, Dict[str, Any]] = {}
    for r in data_b["rows"]:
        lookup_b[_key(r)] = r

    keys_a = set(lookup_a.keys())
    keys_b = set(lookup_b.keys())
    summary["keys_only_in_a"] = len(keys_a - keys_b)
    summary["keys_only_in_b"] = len(keys_b - keys_a)
    summary["common_keys"] = len(keys_a & keys_b)

    # Per-row diffs on common keys
    diffs: List[Dict[str, Any]] = []
    common = cols_a & cols_b
    for k in sorted(keys_a & keys_b):
        if len(diffs) >= head:
            break
        ra, rb = lookup_a[k], lookup_b[k]
        row_diffs: Dict[str, Dict[str, str]] = {}
        for c in sorted(common):
            va, vb = str(ra.get(c, "")), str(rb.get(c, ""))
            if va != vb:
                row_diffs[c] = {"a": va, "b": vb}
        if row_diffs:
            diffs.append({key_column: k, "diffs": row_diffs})

    summary["diff_count"] = len(diffs)
    summary["diffs"] = diffs
    return summary


def lookup_row(
    task_id: str,
    rel_path: str,
    *,
    search_column: str,
    search_value: str,
    match_mode: str = "contains",
    channel_id: Optional[str] = None,
    sheet: Optional[str] = None,
    head: int = 10,
) -> Dict[str, Any]:
    """Search for rows in a CSV/Excel file where ``search_column`` matches ``search_value``.

    match_mode: ``exact``, ``contains`` (default), ``startswith``.
    """
    data = _load_tabular_rows(task_id, rel_path, channel_id=channel_id, sheet=sheet)
    if "error" in data:
        return data
    columns = data["columns"]
    rows = data["rows"]

    if search_column not in columns:
        return {"error": f"column '{search_column}' not found; available: {columns[:20]}"}

    matches: List[Dict[str, Any]] = []
    sv = search_value.lower()
    for row in rows:
        cell = str(row.get(search_column, "")).lower()
        ok = False
        if match_mode == "exact":
            ok = cell == sv
        elif match_mode == "startswith":
            ok = cell.startswith(sv)
        else:  # contains
            ok = sv in cell
        if ok:
            matches.append(row)
            if len(matches) >= head:
                break

    return {
        "rel_path": rel_path,
        "search_column": search_column,
        "search_value": search_value,
        "match_mode": match_mode,
        "total_rows": len(rows),
        "matched": len(matches),
        "rows": matches,
        "head": head,
    }


def verify_summary_stats(
    task_id: str,
    channel_id: str,
    run_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Summarize verify results for a channel run: counts by severity, list warnings/fails.

    Returns pass/warning/fail counts, rule distribution, and the raw warning rows.
    """
    state = StateManager.load_state(task_id)
    if not state or channel_id not in state.channels:
        return {"error": "task or channel not found"}
    ch = state.channels[channel_id]
    if not ch.runs:
        return {"error": "no runs for this channel"}
    if run_id is None:
        run = ch.runs[-1]
    else:
        run = next((r for r in ch.runs if r.run_id == run_id), None)
        if run is None:
            return {"error": "run not found"}

    rows = (run.verify_summary or {}).get("rows", [])
    if not rows:
        return {
            "channel_id": channel_id,
            "run_id": run.run_id,
            "total": 0,
            "counts": {"pass": 0, "warning": 0, "info": 0},
        }

    counts: Dict[str, int] = {}
    by_rule: Dict[str, int] = {}
    for r in rows:
        sev = r.get("severity", "info")
        counts[sev] = counts.get(sev, 0) + 1
        ref = r.get("rule_ref") or "(none)"
        by_rule[ref] = by_rule.get(ref, 0) + 1

    warnings = [r for r in rows if r.get("severity") == "warning"]
    fails = [r for r in rows if r.get("severity") in ("fail", "error")]

    return {
        "channel_id": channel_id,
        "run_id": run.run_id,
        "total": len(rows),
        "counts": counts,
        "by_rule": by_rule,
        "warning_rows": warnings[:20],
        "fail_rows": fails[:20],
    }


# ---------- registry ----------


TOOL_REGISTRY = {
    "list_task_files": list_task_files,
    "read_text": read_text,
    "read_csv": read_csv,
    "read_excel": read_excel,
    "read_verify_summary": read_verify_summary,
    "read_log": read_log,
    "query_rules": query_rules,
    "lookup_password": lookup_password,
    "propose_rule_patch": propose_rule_patch,
    "propose_replace_file": propose_replace_file,
    "mark_row_resolved": mark_row_resolved,
    "filter_table": filter_table,
    "aggregate_table": aggregate_table,
    "compare_files": compare_files,
    "lookup_row": lookup_row,
    "verify_summary_stats": verify_summary_stats,
}
