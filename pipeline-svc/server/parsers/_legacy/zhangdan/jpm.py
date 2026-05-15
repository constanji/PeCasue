import os
import glob
import pandas as pd
import numpy as np
import warnings

"""
    摩根大通银行账单解析工具
"""

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def clean_str(v):
    if pd.isna(v):
        return np.nan
    if isinstance(v, str):
        return v.strip()
    return v

def clean_num(v):
    if pd.isna(v):
        return np.nan
    try:
        return float(v)
    except (ValueError, TypeError):
        return v

def parse_standard(file_path):
    try:
        xl = pd.ExcelFile(file_path)
        sheets = [s for s in xl.sheet_names if s.startswith('Activity Detail')]
        results = []
        
        for sheet in sheets:
            df = xl.parse(sheet, header=None)
            account_no = None
            parsing_items = False
            
            for index, row in df.iterrows():
                val0 = str(row[0]).strip() if pd.notna(row[0]) else ""
                
                if val0.startswith("DEPOSIT ACCOUNT:") or val0.startswith("GROUP ACCOUNT:"):
                    account_no = val0.split(":", 1)[1].strip()
                    continue
                    
                if val0 == "Service\nCode" or val0.startswith("Service"):
                    parsing_items = True
                    continue
                    
                if parsing_items:
                    val1 = str(row[1]).strip() if pd.notna(row[1]) else ""
                    
                    if not val0:
                        continue
                    if not val1:
                        continue
                    if val0.lower().startswith("subtotal") or val0.lower().startswith("total"):
                        continue
                    
                    # Valid row
                    results.append({
                        "DEPOSIT ACCOUNT": account_no,
                        "Service\nCode": clean_str(row[0]),
                        "Description": clean_str(row[1]),
                        "Volume": clean_num(row[3]),
                        "Curr": clean_str(row[4]),
                        "Unit\nPrice": clean_num(row[5]),
                        "Price\nID": clean_str(row[6]),
                        "Charge for\nService": clean_num(row[7]),
                        "Source": os.path.basename(file_path)
                    })
        return results
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        return []

def parse_802(file_path):
    try:
        xl = pd.ExcelFile(file_path)
        if 'Activity Summary' not in xl.sheet_names:
            print(f"Warning: 'Activity Summary' sheet not found in {file_path}")
            return []
            
        df = xl.parse('Activity Summary', header=None)
        account_no = None
        parsing_items = False
        results = []
        
        for index, row in df.iterrows():
            val0 = str(row[0]).strip() if pd.notna(row[0]) else ""
            
            if val0.startswith("DEPOSIT ACCOUNT:") or val0.startswith("GROUP ACCOUNT:"):
                account_no = val0.split(":", 1)[1].strip()
                continue
                
            if val0 == "PRODUCT LINE AND DESCRIPTION":
                parsing_items = True
                continue
                
            if parsing_items:
                val1 = str(row[1]).strip() if pd.notna(row[1]) else ""
                
                if not val0:
                    continue
                if not val1:
                    continue
                if val0.lower().startswith("total") or val0.lower().startswith("subtotal"):
                    continue
                    
                # Valid row
                results.append({
                    "DEPOSIT ACCOUNT": account_no,
                    "Service\nCode": clean_str(row[1]),
                    "Description": clean_str(row[0]),
                    "Volume": clean_num(row[4]),
                    "Curr": "USD",
                    "Unit\nPrice": clean_num(row[3]),
                    "Price\nID": clean_str(row[2]),
                    "Charge for\nService": clean_num(row[5]),
                    "Source": os.path.basename(file_path)
                })
        return results
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")
        return []

def parse_earnings_allowance(file_path):
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = None
        for s in xl.sheet_names:
            if 'Balance and Compensation' in s:
                sheet_name = s
                break
                
        if not sheet_name:
            return []
            
        account_no = None
        for s in xl.sheet_names:
            df_temp = xl.parse(s, header=None)
            for idx, row in df_temp.iterrows():
                val0 = str(row[0]).strip() if pd.notna(row[0]) else ""
                if val0.startswith("DEPOSIT ACCOUNT:") or val0.startswith("GROUP ACCOUNT:"):
                    account_no = val0.split(":", 1)[1].strip()
                    break
            if account_no:
                break
                
        if not account_no or account_no.replace(" ", "") != "0802000000000623296537":
            return []
            
        df = xl.parse(sheet_name, header=None)
        results = []
        
        desc_col = -1
        comp_col = -1
        header_idx = -1
        
        for idx, row in df.iterrows():
            row_strs = [str(x).strip().upper().replace('\n', ' ') for x in row if pd.notna(x)]
            has_desc = any('DESCRIPTION' in x for x in row_strs)
            has_comp = any('COMPENSATION INFORMATION' in x for x in row_strs)
            
            if has_desc and has_comp:
                header_idx = idx
                for c_idx, val in row.items():
                    if pd.notna(val):
                        val_str = str(val).strip().upper().replace('\n', ' ')
                        if 'DESCRIPTION' in val_str:
                            desc_col = c_idx
                        elif 'COMPENSATION INFORMATION' in val_str:
                            comp_col = c_idx
                break
                
        if header_idx != -1 and desc_col != -1 and comp_col != -1:
            for idx, row in df.loc[header_idx+1:].iterrows():
                if len(row) > max(desc_col, comp_col):
                    desc_val = str(row[desc_col]).strip() if pd.notna(row[desc_col]) else ""
                    if desc_val.upper() == 'EARNINGS ALLOWANCE':
                        comp_val = clean_num(row[comp_col])
                        if pd.notna(comp_val):
                            try:
                                charge_val = -abs(float(comp_val))
                                results.append({
                                    "DEPOSIT ACCOUNT": account_no,
                                    "Description": "EARNINGS ALLOWANCE",
                                    "Curr": "USD",
                                    "Charge for\nService": charge_val,
                                    "Source": os.path.basename(file_path)
                                })
                                return results
                            except:
                                pass
                            
        # Fallback: 如果没有找到表头，直接找最后面的数值
        for idx, row in df.iterrows():
            for c_idx, val in row.items():
                if pd.notna(val) and str(val).strip().upper() == 'EARNINGS ALLOWANCE':
                    # 从后往前找，因为 compensation 通常在最右侧的列，而中间可能是 balance（通常是0）
                    for right_idx in range(len(row) - 1, c_idx, -1):
                        comp_val = clean_num(row[right_idx])
                        if pd.notna(comp_val) and isinstance(comp_val, float):
                            charge_val = -abs(float(comp_val))
                            results.append({
                                "DEPOSIT ACCOUNT": account_no,
                                "Description": "EARNINGS ALLOWANCE",
                                "Curr": "USD",
                                "Charge for\nService": charge_val,
                                "Source": os.path.basename(file_path)
                            })
                            return results
                            
        return results
    except Exception as e:
        print(f"Error parsing EARNINGS ALLOWANCE in {file_path}: {e}")
        return []

def is_valid_filename(filename):
    """
    根据给定的规则判断文件名是否需要提取
    """
    valid_parts = [
        "655-0023", "655-0024", "655-0025", "655-0026", "655-0027",
        "655-0028", "655-0029", "655-0030", "655-0031", "655-0408",
        "655-5654", "655-7524", "655-7537", "655-7541", "655-7716",
        "671-5770", "671-5775", "671-6849", "671-6851", "671-8959",
        "802-6537-JPM-Billing-L-0003", "6300150684", "6331724044"
    ]
    
    # 所有 883- 前缀的特殊处理，因为文件名中可能在其他位置，或者要求文件名以883-开头/包含883-
    # 根据用户描述 "所有883-前缀"，这里使用 in 包含判断
    if "883-" in filename:
        return True
        
    for part in valid_parts:
        if part in filename:
            return True
            
    return False

def process_all(input_dir, output_file):
    print(f"开始处理目录: {input_dir}")
    all_results = []
    
    # Recursively find all xlsx files, ignoring temp files
    excel_files = glob.glob(os.path.join(input_dir, '**', '*.xlsx'), recursive=True)
    excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]
    
    # 根据规则过滤文件
    valid_files = [f for f in excel_files if is_valid_filename(os.path.basename(f))]
    
    print(f"找到 {len(excel_files)} 个 Excel 文件，其中符合规则的文件有 {len(valid_files)} 个。")
    
    for file_path in valid_files:
        filename = os.path.basename(file_path)
        if '-802-' in filename:
            res = parse_802(file_path)
        else:
            res = parse_standard(file_path)
        
        if res:
            all_results.extend(res)
            
        ea_res = parse_earnings_allowance(file_path)
        if ea_res:
            all_results.extend(ea_res)
            
    df = pd.DataFrame(all_results)
    
    if not df.empty:
        # 补齐目标列
        columns_order = ['DEPOSIT ACCOUNT', 'Service\nCode', 'Description', 
                         'Volume', 'Curr', 'Unit\nPrice', 'Price\nID', 'Charge for\nService', 'Source']
        for col in columns_order:
            if col not in df.columns:
                df[col] = np.nan
        df = df[columns_order]
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                
                # 获取最大行和列的字母表示 (如 A1:G62)
                import openpyxl
                max_col = openpyxl.utils.get_column_letter(worksheet.max_column)
                max_row = worksheet.max_row
                
                # 添加自动筛选功能
                worksheet.auto_filter.ref = f"A1:{max_col}{max_row}"
                
            print(f"\n提取完成！共处理了 {len(valid_files)} 个符合规则的文件，提取了 {len(df)} 条明细。")
            print(f"结果已保存至: {output_file}")
        except Exception as e:
            print(f"保存 Excel 时出错: {e}")
    else:
        print("\n未提取到任何明细数据。")

if __name__ == "__main__":
    input_directory = r"E:\2月成本分摊模拟\202602\2026.02账单\JPM账单"
    output_excel_path = r"e:\Desktop\demo\2026.2JPM.xlsx"
    process_all(input_directory, output_excel_path)
