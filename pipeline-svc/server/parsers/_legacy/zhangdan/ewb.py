"""
East West Bank（EWB）Account Analysis PDF 明细提取 — 单文件完整流程。

合并自 ewb/ 目录下原 ewb_statement 包与 extract_ewb_pdf 入口逻辑；
原 cli 依赖的 bill.bill_output（ewb_to_bill / empty_bill）在本项目中不存在，
此处提供等价实现：非 --raw 时用 period 覆盖 Date 列（对应 all.py 映射中的「入账期间」）。
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

# ---------------------------------------------------------------------------
# constants（原 ewb_statement/constants.py）
# ---------------------------------------------------------------------------

OUTPUT_COLUMNS = [
    "Reference",
    "Date",
    "DESCRIPTION",
    "VOLUME",
    "UNIT PRICE",
    "TOTAL CHARGE",
    "TOTAL CHARGE CCY",
    "Waiver",
]

META_COLUMNS = ["source_pdf", "source_stem", "page"]

VOLUME_LINE_RE = re.compile(r"^[\d,]+$")

ACCOUNT_NUMBER_RE = re.compile(
    r"Account Number\s*\n?\s*(\d+)", re.MULTILINE | re.IGNORECASE
)
STATEMENT_DATE_RE = re.compile(
    r"Statement Date\s*\n?\s*([^\n]+)", re.MULTILINE | re.IGNORECASE
)
STATEMENT_NO_IN_STEM_RE = re.compile(r"#(\d+)")

DEFAULT_CCY = "USD"

# ---------------------------------------------------------------------------
# text（原 ewb_statement/text.py）
# ---------------------------------------------------------------------------

_HISTORICAL_SUMMARY_MARKER = "Historical Summary"

_SERVICE_DETAIL_END_MARKERS = (
    "Total Analyzed Results",
    "Total Analyzed Service Charges This Statement",
    "Total Analyzed Fees",
)


def trim_before_historical_summary(text: str) -> str:
    if _HISTORICAL_SUMMARY_MARKER not in text:
        return text
    idx = text.index(_HISTORICAL_SUMMARY_MARKER)
    return text[:idx].rstrip()


def trim_to_service_detail_body(text: str) -> str:
    lines = text.splitlines()
    stripped = [ln.strip() for ln in lines]

    start = 0
    for i, line in enumerate(stripped):
        if "Charged to Account" in line or "Transferred to Composite Account" in line:
            start = i + 1
            break
    work = stripped[start:]

    end = len(work)
    for i, line in enumerate(work):
        if line in _SERVICE_DETAIL_END_MARKERS:
            end = i
            break
    work = work[:end]

    return "\n".join(work)


def trim_before_footer(lines: list[str]) -> list[str]:
    out: list[str] = []
    for ln in lines:
        s = ln.strip()
        if s == "Account Number" or re.match(r"^Account Number\s+\d", s, re.I):
            break
        if s.startswith("$0.00 in Investable Balances Offset"):
            break
        out.append(ln)
    return out


def normalize_lines(text: str) -> list[str]:
    return [ln.strip() for ln in text.splitlines() if ln.strip()]

# ---------------------------------------------------------------------------
# detect（原 ewb_statement/detect.py）
# ---------------------------------------------------------------------------

_HEADER_LEN = 3500

_SERVICE_DETAIL_CONTINUED_RE = re.compile(
    r"Service\s+Detail\s*[-–]\s*Continued", re.IGNORECASE
)
_NEW_SERVICE_DETAIL_RE = re.compile(
    r"Service\s+Detail(?!-Continued)", re.IGNORECASE
)


def _header_text(text: str) -> str:
    return text[:_HEADER_LEN] if len(text) > _HEADER_LEN else text


def is_service_detail_continued_header(text: str) -> bool:
    return bool(_SERVICE_DETAIL_CONTINUED_RE.search(_header_text(text)))


def is_new_service_detail_header(text: str) -> bool:
    h = _header_text(text)
    if is_service_detail_continued_header(text):
        return False
    return bool(_NEW_SERVICE_DETAIL_RE.search(h))


def looks_like_ewb_service_page(text: str) -> bool:
    u = text.upper()
    return (
        "EAST WEST BANK" in u
        and "SERVICE DESCRIPTION" in u
        and "TOTAL FEE" in u
        and "UNIT PRICE" in u
    )

# ---------------------------------------------------------------------------
# money（原 ewb_statement/money.py）
# ---------------------------------------------------------------------------


def parse_money_token(s: str) -> str | None:
    s = s.strip()
    if not s:
        return None
    neg = "(" in s and ")" in s
    s = s.replace("$", "").replace(",", "").strip()
    s = re.sub(r"^\(\s*", "", s)
    s = re.sub(r"\s*\)$", "", s)
    if not s:
        return None
    try:
        v = float(s)
        if neg:
            v = -abs(v)
    except ValueError:
        return None
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return format(v, "f").rstrip("0").rstrip(".")

# ---------------------------------------------------------------------------
# filters（原 ewb_statement/filters.py）
# ---------------------------------------------------------------------------


def is_waived_line(s: str) -> bool:
    return s.strip().casefold() == "waived"


def should_skip_desc_line(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    sl = s.lower()
    if s.startswith("$") or s.startswith("("):
        return True
    if VOLUME_LINE_RE.match(s.replace(",", "")):
        return True
    if s.startswith("Total"):
        return True
    if is_waived_line(s):
        return True
    if sl.startswith("less "):
        return True
    skip_substrings = (
        "account number",
        "statement date",
        "settlement period",
        "relationship",
        "east west bank",
        "pingpong",
        "pasadena",
        "fashion island",
        "service description",
        "account analysis",
        "earnings credit",
        "less total analyzed",
        "the total deficit",
        "average ledger",
        "average float",
        "average collected",
        "collected balance",
        "investable balance",
        "reserve requirement",
        "excess/(deficit)",
        "less balance",
        "total analyzed results",
        "total analyzed service charges",
        "historical summary",
    )
    if any(x in sl for x in skip_substrings):
        return True
    if "balance required" in sl and sl != "balance required":
        return True
    if re.match(r"^page\s+\d+", sl):
        return True
    return False

# ---------------------------------------------------------------------------
# meta（原 ewb_statement/meta.py）
# ---------------------------------------------------------------------------


def extract_account_number(text: str) -> str:
    m = ACCOUNT_NUMBER_RE.search(text)
    return m.group(1) if m else ""


def extract_statement_date(text: str) -> str:
    m = STATEMENT_DATE_RE.search(text)
    return m.group(1).strip() if m else ""


def build_reference(account: str, pdf_stem: str) -> str:
    stmt_m = STATEMENT_NO_IN_STEM_RE.search(pdf_stem)
    stmt_no = stmt_m.group(1) if stmt_m else ""
    if stmt_no and account:
        return f"{stmt_no}-{account}"
    if account:
        return account
    return stmt_no


def stem_from_path(path: Path) -> str:
    return path.stem

# ---------------------------------------------------------------------------
# parser（原 ewb_statement/parser.py）
# ---------------------------------------------------------------------------


def _is_balance_required(desc: str) -> bool:
    return re.sub(r"\s+", " ", desc.strip()).casefold() == "balance required"


def extract_service_rows_from_page(
    page_text: str, meta_text: str | None, pdf_stem: str
) -> list[dict[str, str]]:
    head = meta_text if meta_text is not None else page_text
    account = extract_account_number(head)
    ref = build_reference(account, pdf_stem)
    date = extract_statement_date(head)

    raw_lines = normalize_lines(page_text)
    lines = trim_before_footer(raw_lines)
    out: list[dict[str, str]] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if should_skip_desc_line(line):
            i += 1
            continue
        if i + 1 >= len(lines):
            i += 1
            continue
        vol_line = lines[i + 1]
        if not VOLUME_LINE_RE.match(vol_line.replace(",", "")):
            i += 1
            continue
        desc = line
        vol = vol_line.replace(",", "")
        i += 2
        money_lines: list[str] = []
        waived_seen = False
        while i < len(lines):
            s = lines[i]
            st = s.strip()
            if is_waived_line(st):
                waived_seen = True
                i += 1
                break
            if st.startswith("$") or (st.startswith("(") and "$" in st):
                money_lines.append(s)
                i += 1
                if len(money_lines) >= 2:
                    break
                continue
            break
        if i < len(lines) and is_waived_line(lines[i]):
            waived_seen = True
            i += 1
        unit_s, total_s = "", ""
        if len(money_lines) == 2:
            unit_s = parse_money_token(money_lines[0]) or ""
            total_s = parse_money_token(money_lines[1]) or ""
        elif len(money_lines) == 1:
            total_s = parse_money_token(money_lines[0]) or ""
        else:
            if _is_balance_required(desc) and waived_seen:
                total_s = "0"
            else:
                continue
        if waived_seen:
            total_s = "0"
        if not desc or not total_s:
            continue
        rec = {k: "" for k in OUTPUT_COLUMNS}
        rec["Reference"] = ref
        rec["Date"] = date
        rec["DESCRIPTION"] = desc
        rec["VOLUME"] = vol
        rec["UNIT PRICE"] = unit_s
        rec["TOTAL CHARGE"] = total_s
        rec["TOTAL CHARGE CCY"] = DEFAULT_CCY
        if waived_seen:
            rec["Waiver"] = "Waived"
        out.append(rec)
    return out


def attach_meta(
    rows: list[dict[str, str]],
    source_pdf: str,
    source_stem: str,
    page: int,
) -> list[dict[str, Any]]:
    meta = {"source_pdf": source_pdf, "source_stem": source_stem, "page": page}
    return [{**meta, **r} for r in rows]

# ---------------------------------------------------------------------------
# extract_pdf（原 ewb_statement/extract_pdf.py）
# ---------------------------------------------------------------------------


def extract_from_pdf(path: Path) -> list[dict[str, Any]]:
    path = path.expanduser().resolve()
    stem = stem_from_path(path)
    try:
        import fitz
    except ImportError as e:
        raise RuntimeError("需要安装 pymupdf：pip install pymupdf") from e

    out: list[dict[str, Any]] = []
    doc = fitz.open(str(path))
    in_first_block = False
    try:
        for pi in range(len(doc)):
            raw = doc[pi].get_text("text") or ""
            hist = trim_before_historical_summary(raw)
            if not looks_like_ewb_service_page(hist):
                continue
            is_new = is_new_service_detail_header(hist)
            if not in_first_block:
                if not is_new:
                    continue
                in_first_block = True
            else:
                if is_new:
                    break
            body = trim_to_service_detail_body(hist)
            rows = extract_service_rows_from_page(body, hist, stem)
            out.extend(
                attach_meta(rows, str(path), stem, pi + 1)
            )
    finally:
        doc.close()
    return out

# ---------------------------------------------------------------------------
# output_utils（原 ewb_statement/output_utils.py）
# ---------------------------------------------------------------------------


def parse_total_charge_value(rec: dict[str, Any]) -> float | None:
    raw = str(rec.get("TOTAL CHARGE", "") or "").strip().replace(",", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def drop_zero_total_charge_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in rows:
        v = parse_total_charge_value(r)
        desc = str(r.get("DESCRIPTION", "") or "").strip().lower()
        waiver = str(r.get("Waiver", "") or "").strip().casefold()
        keep_zero = desc == "balance required" or waiver == "waived"
        if v is not None and v == 0.0 and not keep_zero:
            continue
        out.append(r)
    return out


def dedupe_by_business_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_file: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_file.setdefault(r.get("source_pdf", ""), []).append(r)
    out: list[dict[str, Any]] = []
    for _path, group in by_file.items():
        seen: set[tuple[str, ...]] = set()
        for r in group:
            key = tuple(str(r.get(k, "") or "") for k in OUTPUT_COLUMNS)
            if key in seen:
                continue
            seen.add(key)
            out.append(r)
    return out

# ---------------------------------------------------------------------------
# 原 bill.bill_output 等价（仅 EWB 汇总列，供 CLI 非 --raw 与 empty 使用）
# ---------------------------------------------------------------------------


def midout_dir() -> Path:
    return Path(__file__).resolve().parent / "midout"


def _export_column_order() -> list[str]:
    omit: set[str] = {"source_pdf", "page"}
    base_cols = META_COLUMNS + OUTPUT_COLUMNS
    out_cols = [c for c in base_cols if c not in omit]
    out_cols.append("source")
    return out_cols


def empty_bill() -> pd.DataFrame:
    return pd.DataFrame(columns=_export_column_order())


def ewb_to_bill(df_raw: pd.DataFrame, period: str) -> pd.DataFrame:
    base_cols = META_COLUMNS + OUTPUT_COLUMNS
    df = df_raw.copy()
    for c in base_cols:
        if c not in df.columns:
            df[c] = ""
    df["source"] = df["source_pdf"].map(lambda p: Path(str(p)).name if p else "")
    df["Date"] = str(period)
    out_cols = _export_column_order()
    return df[out_cols].copy()

# ---------------------------------------------------------------------------
# CLI（原 ewb_statement/cli.py + ewb/extract_ewb_pdf.py）
# ---------------------------------------------------------------------------


def iter_pdfs(root: Path, recursive: bool) -> Iterable[Path]:
    if root.is_file() and root.suffix.lower() == ".pdf":
        yield root
        return
    pattern = "**/*.pdf" if recursive else "*.pdf"
    yield from sorted(root.glob(pattern))


def cli_main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="EWB Account Analysis：仅 Service Detail / Service Detail-Continued 明细，输出 CSV/Excel",
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help="PDF 文件或目录",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出 .csv 或 .xlsx，默认 <项目>/midout/ewb_analysis_extracted.csv",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="目录时递归子文件夹",
    )
    parser.add_argument(
        "--omit-columns",
        type=str,
        default="",
        help="额外排除的列，逗号分隔（如 source_stem）",
    )
    parser.add_argument(
        "--with-source-locations",
        action="store_true",
        help="保留 source_pdf、page（默认不输出这两列）",
    )
    parser.add_argument(
        "--keep-zero",
        action="store_true",
        help="保留 TOTAL CHARGE 为 0 的行（默认会剔除）",
    )
    parser.add_argument(
        "--period",
        type=str,
        default="202602",
        help="非 --raw 时写入 Date 列（入账期间）；默认 202602",
    )
    parser.add_argument(
        "--raw",
        action="store_true",
        help="输出 PDF 抽取原始列；默认输出与 ewb_to_bill 一致的列",
    )
    args = parser.parse_args(argv)
    input_path = args.input_path.expanduser().resolve()
    out = args.output
    if out is None:
        out = midout_dir() / "ewb_analysis_extracted.csv"
    else:
        out = out.expanduser().resolve()

    pdfs = list(iter_pdfs(input_path, args.recursive))
    if not pdfs:
        print("未找到 PDF。", file=sys.stderr)
        sys.exit(1)

    records: list = []
    for p in pdfs:
        try:
            records.extend(extract_from_pdf(p))
        except Exception as e:
            print(f"[跳过] {p}: {e}", file=sys.stderr)

    records = dedupe_by_business_columns(records)
    if not args.keep_zero:
        records = drop_zero_total_charge_rows(records)

    omit: set[str] = {x.strip() for x in args.omit_columns.split(",") if x.strip()}
    if not args.with_source_locations:
        omit.update(("source_pdf", "page"))
    base_cols = META_COLUMNS + OUTPUT_COLUMNS
    out_cols = [c for c in base_cols if c not in omit]
    if "source" not in omit:
        out_cols.append("source")

    if not records:
        df = pd.DataFrame(columns=out_cols) if args.raw else empty_bill()
    else:
        df_raw = pd.DataFrame(records)
        for c in base_cols:
            if c not in df_raw.columns:
                df_raw[c] = ""
        if "source" not in omit:
            df_raw["source"] = df_raw["source_pdf"].map(
                lambda p: Path(str(p)).name if p else ""
            )
        if args.raw:
            df = df_raw[out_cols]
        else:
            df = ewb_to_bill(df_raw, args.period)
            df = df[[c for c in df.columns if c not in omit]]

    out.parent.mkdir(parents=True, exist_ok=True)
    if out.suffix.lower() == ".xlsx":
        df.to_excel(out, index=False)
    else:
        df.to_csv(out, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)

    print(f"PDF 数: {len(pdfs)}，行数: {len(df)}，输出: {out}")

# ---------------------------------------------------------------------------
# 与 all.py / main.py 对接的入口（对齐 barclays.py：extract_pdf_data + main）
# ---------------------------------------------------------------------------


def extract_pdf_data(pdf_path: str | os.PathLike[str]) -> list[dict[str, Any]]:
    """
    单份 PDF 解析（与 barclays.extract_pdf_data 同名同角色）。
    失败时打印错误并返回空列表，不向外抛异常。
    """
    try:
        return extract_from_pdf(Path(pdf_path).expanduser().resolve())
    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")
        return []


def main(input_folder: str, output_excel: str) -> None:
    input_folder = str(Path(input_folder).expanduser().resolve())
    pdf_files = sorted(glob.glob(os.path.join(input_folder, "*.pdf")))
    if not pdf_files:
        print(f"未找到 PDF 文件: {input_folder}")
        return

    print(f"找到 {len(pdf_files)} 个 PDF 文件，正在开始解析...")

    all_records: list[dict[str, Any]] = []
    for pdf_path in pdf_files:
        all_records.extend(extract_pdf_data(pdf_path))

    all_records = dedupe_by_business_columns(all_records)
    all_records = drop_zero_total_charge_rows(all_records)

    if not all_records:
        print("未能提取到任何有效数据。")
        return

    omit = {"source_pdf", "page"}
    base_cols = META_COLUMNS + OUTPUT_COLUMNS
    out_cols = [c for c in base_cols if c not in omit]
    out_cols.append("source")

    df = pd.DataFrame(all_records)
    for c in base_cols:
        if c not in df.columns:
            df[c] = ""
    df["source"] = df["source_pdf"].map(lambda p: Path(str(p)).name if p else "")
    df = df[out_cols]

    output_excel = str(Path(output_excel).expanduser().resolve())
    Path(output_excel).parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]
            import openpyxl
            from openpyxl.utils import get_column_letter

            max_col = get_column_letter(worksheet.max_column)
            max_row = worksheet.max_row
            worksheet.auto_filter.ref = f"A1:{max_col}{max_row}"

        print(f"成功提取了 {len(df)} 条明细记录，并已保存至: {output_excel}")
    except Exception as e:
        print(f"保存 Excel 时出错: {e}")


if __name__ == "__main__":
    input_folder = r"E:\2月成本分摊模拟\202602\2026.02账单\EWB账单"
    output_excel = r"e:\Desktop\demo\ewb账单汇总.xlsx"
    main(input_folder, output_excel)
