import argparse
import glob
import os
import re
import sys
from pathlib import Path

import pandas as pd
import PyPDF2

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

"""
    巴克莱银行账单解析工具
"""

# 新版式/不同渲染下，锚点行不一定与历史 PDF 完全一致（空格、全角括号、换行）
_RE_CHARGES_START = re.compile(
    r"^\s*\[?\s*Charges\s+billed\s+this\s+period\s*\]?\s*$",
    re.IGNORECASE,
)


def _extract_text_pypdf2(pdf_path: str) -> str:
    with open(pdf_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        parts: list[str] = []
        for page in reader.pages:
            parts.append(page.extract_text() or "")
        return "\n".join(parts)


def _extract_text_pdfplumber(pdf_path: str) -> str:
    if pdfplumber is None:
        return ""
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)


def _extract_pdf_text(pdf_path: str) -> tuple[str, str]:
    """
    返回 (全文, 来源说明)。
    PyPDF2 常抽不出某些新版式；文字过少时换 pdfplumber 再试。
    """
    try:
        text = _extract_text_pypdf2(pdf_path)
    except Exception as e:
        print(f"PyPDF2 读取 {os.path.basename(pdf_path)}: {e}")
        text = ""
    src = "PyPDF2"
    if len(text.strip()) < 80 and pdfplumber is not None:
        alt = _extract_text_pdfplumber(pdf_path)
        if len(alt.strip()) > len(text.strip()):
            text = alt
            src = "pdfplumber"
    return text, src


def _charges_section_start(line: str) -> bool:
    s = line.strip()
    if s == "[Charges billed this period]":
        return True
    return bool(_RE_CHARGES_START.match(s))


def extract_pdf_data(pdf_path: str) -> list[dict]:
    try:
        text, text_src = _extract_pdf_text(pdf_path)
    except Exception as e:
        print(f"Error reading {pdf_path}: {e}")
        return []

    if text_src == "pdfplumber":
        print(
            f"  提示: {os.path.basename(pdf_path)} 已用 pdfplumber 重抽文字（本页 PyPDF2 结果过短）。"
        )

    if not text or len(text.strip()) < 20:
        print(
            f"  提示: {os.path.basename(pdf_path)} 抽到的文字极少（<20 字），"
            "可能是扫描件/无文字层，需 OCR 或另存为文本型 PDF。"
        )
        return []
        
    date_match = re.search(r"Date:\s*(.+)", text)
    ref_match = re.search(r"Reference:\s*(\d+)", text)
    
    date = date_match.group(1).strip() if date_match else ""
    ref = ref_match.group(1).strip() if ref_match else ""
    
    # 格式化日期为 YYYY-MM-DD
    if date:
        try:
            date_obj = pd.to_datetime(date)
            date_str = date_obj.strftime("%Y-%m-%d")
        except:
            date_str = date
    else:
        date_str = ""
    
    lines = text.split('\n')
    in_charges = False
    charges_header_found = False

    results = []
    for line in lines:
        line = line.strip()
        
        # 遇到总计即停止，防止读取后面的分账户明细产生重复
        clean_line = line.replace(" ", "")
        if "TotalChargeExcludingTaxes" in clean_line or "TotalChargesDue" in clean_line:
            break
            
        if _charges_section_start(line):
            in_charges = True
            charges_header_found = True
            continue
        if in_charges and line.strip().lower().startswith("total for"):
            in_charges = False
            continue
            
        if in_charges:
            parts = line.split()
            if len(parts) < 3:
                continue
                
            ccy = parts[-1]
            total_charge = parts[-2]
            
            total_charge_clean = total_charge.replace(',', '')
            if not re.match(r"^-?\d+(?:\.\d+)?$", total_charge_clean):
                continue
                
            # 过滤掉 0.0 的费用
            if float(total_charge_clean) == 0.0:
                continue
                
            volume = None
            unit_price = None
            desc_parts = parts[:-2]
            
            if len(desc_parts) > 0 and re.match(r"^-?[\d,]+(?:\.\d+)?$", desc_parts[-1]):
                val1 = desc_parts.pop().replace(',', '')
                if len(desc_parts) > 0 and re.match(r"^-?[\d,]+(?:\.\d+)?$", desc_parts[-1]):
                    val2 = desc_parts.pop().replace(',', '')
                    volume = float(val2)
                    unit_price = float(val1)
                else:
                    volume = float(val1)
            
            description = " ".join(desc_parts)
            results.append({
                "Reference": ref,
                "Date": date_str,
                "DESCRIPTION": description,
                "VOLUME": volume,
                "UNIT PRICE": unit_price,
                "TOTAL CHARGE": float(total_charge_clean),
                "TOTAL CHARGE CCY": ccy,
                "Source": os.path.basename(pdf_path)
            })

    if not results and text:
        low = text.lower()
        if not charges_header_found:
            if "charge" in low and "billed" in low:
                print(
                    f"  提示: {os.path.basename(pdf_path)} 含 charge/billed 字样但未匹配到"
                    "「Charges billed this period」标题行，可能被拆行或标点不同。"
                )
            else:
                print(
                    f"  提示: {os.path.basename(pdf_path)} 未出现可识别的计费段标题，"
                    "银行可能已改版式，或 PDF 为扫描件无文字层。"
                )
        else:
            print(
                f"  提示: {os.path.basename(pdf_path)} 已识别计费段但无有效明细行"
                "（常见：金额均为 0、或列格式与脚本假定不一致）。"
            )

    return results

def main(input_folder, output_excel):
    pdf_files = glob.glob(os.path.join(input_folder, "*.pdf"))
    if not pdf_files:
        print(f"未找到 PDF 文件: {input_folder}")
        return
        
    print(f"找到 {len(pdf_files)} 个 PDF 文件，正在开始解析...")
    
    all_records = []
    for pdf_path in pdf_files:
        records = extract_pdf_data(pdf_path)
        all_records.extend(records)
        
    if not all_records:
        print("未能提取到任何有效数据。")
        return
        
    df = pd.DataFrame(all_records)
    
    # 确保列名顺序
    columns = ["Reference", "Date", "DESCRIPTION", "VOLUME", "UNIT PRICE", "TOTAL CHARGE", "TOTAL CHARGE CCY", "Source"]
    df = df[columns]
    
    # 将 Reference 保存为文本，避免前导 0 被吃掉
    # 但 Excel 可能还会将其识别为数字。使用 openpyxl 写入字符串格式更好。
    # 为了保证格式与样例一致，直接保存即可
    
    # 输出到 Excel，并为表头添加自动筛选功能
    try:
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # 获取最大行和列的字母表示 (如 A1:G62)
            import openpyxl
            max_col = openpyxl.utils.get_column_letter(worksheet.max_column)
            max_row = worksheet.max_row
            
            # 添加自动筛选功能
            worksheet.auto_filter.ref = f"A1:{max_col}{max_row}"
            
        print(f"成功提取了 {len(df)} 条明细记录，并已保存至: {output_excel}")
    except Exception as e:
        print(f"保存 Excel 时出错: {e}")

if __name__ == "__main__":
    input_folder = r"e:\Desktop\demo\barclays账单"
    output_excel = r"e:\Desktop\demo\barclays账单汇总.xlsx"
    main(input_folder, output_excel)
