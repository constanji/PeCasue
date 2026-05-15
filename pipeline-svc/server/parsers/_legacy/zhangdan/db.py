from __future__ import annotations

import re
import os
import glob
import zipfile
import pandas as pd
import pdfplumber

"""
    德意志银行账单解析工具
"""

def parse_account_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
            
    account_no_match = re.search(r'Account No\s*:\s*([\w\-]+)', text)
    account_no = account_no_match.group(1) if account_no_match else "UNKNOWN"
    
    lines = text.split('\n')
    parsing_items = False
    current_service_desc = ""
    results = []
    
    for line in lines:
        if "Item TariffCCY Amount Avg Per" in line or "Count Item" in line:
            parsing_items = True
            continue
        if "For settlement details" in line or "Tariff Composition" in line:
            parsing_items = False
            continue
            
        if parsing_items:
            line_str = line.strip()
            if not line_str:
                continue
                
            # 尝试解析当前行是否是费用明细行（至少需要有 Item Count, Ccy, Amount, Avg Per Item 这几个数字部分）
            parts = line_str.split()
            is_detail_row = False
            
            if len(parts) >= 4:
                try:
                    # 检查最后两个元素是否都能转为浮点数 (Amount 和 Avg Per Item)
                    float(parts[-1].replace(',', ''))
                    float(parts[-2].replace(',', ''))
                    is_detail_row = True
                except ValueError:
                    is_detail_row = False

            if not is_detail_row:
                # 如果不是费用明细行，则这一行就是大标题 (Service Description)
                current_service_desc = line_str
            else:
                try:
                    avg_per_item = float(parts[-1].replace(',', ''))
                    amount = float(parts[-2].replace(',', ''))
                    ccy = parts[-3]
                    
                    if parts[-4] in ['F', 'M', 'P', 'A', 'R', 'S']:
                        item_count = float(parts[-5].replace(',', ''))
                        prod_desc = " ".join(parts[:-5])
                    else:
                        item_count = float(parts[-4].replace(',', ''))
                        prod_desc = " ".join(parts[:-4])
                        
                    if amount > 0:
                        results.append({
                            'Service Type': 'Account Details',
                            'Account No': account_no,
                            'Product Description': prod_desc,
                            'Item Count': item_count,
                            'Tariff CCY': ccy,
                            'Amount': amount,
                            'Avg Per Item': avg_per_item,
                            'Source': os.path.basename(pdf_path)
                        })
                except ValueError:
                    pass
    return results

def parse_summary_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
            
    lines = text.split('\n')
    parsing_periodic = False
    parsing_reimbursement = False
    results = []
    
    # 第一遍遍历提取 Periodic Fees
    for line in lines:
        line_str = line.strip()
        if "Periodic Fee Currency" in line_str or "Periodic Fees" in line_str:
            if "Ccy Amount" in line_str or "Periodic Fees" == line_str:
                parsing_periodic = True
            continue
        if "In reimbursement the following accounts" in line_str:
            parsing_periodic = False
            continue
            
        if parsing_periodic:
            if not line_str or line_str == "Ccy Amount":
                continue
            
            parts = line_str.split()
            if len(parts) >= 3:
                try:
                    amount = float(parts[-1].replace(',', ''))
                    ccy = parts[-2]
                    prod_desc = " ".join(parts[:-2])
                    if amount > 0:
                        results.append({
                            'Service Type': 'Periodic Fees',
                            'Account No': 'summary',
                            'Product Description': prod_desc,
                            'Item Count': None,
                            'Tariff CCY': ccy,
                            'Amount': amount,
                            'Avg Per Item': None,
                            'Source': os.path.basename(pdf_path)
                        })
                except ValueError:
                    pass

    # 第二遍遍历提取实际被扣款的账号 (In reimbursement the following accounts will be debited on)
    reimbursement_account = 'summary'
    for line in lines:
        line_str = line.strip()
        if "In reimbursement the following accounts" in line_str:
            parsing_reimbursement = True
            continue
            
        if parsing_reimbursement:
            if not line_str or "Debit Currency" in line_str or "Ccy Amount Due" in line_str:
                continue
            
            # 停止条件：遇到可能标识该区块结束的文本
            if "For settlement details" in line_str or "Summary" in line_str:
                parsing_reimbursement = False
                continue

            # 例如: PING PONG GLOBAL HOLDINGS LIMITED 0037341001-HKD HKD 295.00
            parts = line_str.split()
            if len(parts) >= 4:
                # 倒数第三部分通常是 账号-币种，如 0037341001-HKD
                possible_account = parts[-3]
                if '-' in possible_account:
                    # 保留完整账号（包括后缀），如 0037341001-HKD
                    reimbursement_account = possible_account
                    break
                    
    # 更新之前提取结果中的账号
    for res in results:
        if res['Account No'] == 'summary':
            res['Account No'] = reimbursement_account

    return results


def _check_zip_needs_password(zip_path: str) -> str | None:
    """若 ZIP 需密码或无法打开，返回说明字符串；否则 None。"""
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            if not zf.namelist():
                return "空 ZIP 文件"
            for info in zf.infolist():
                if info.filename.endswith("/"):
                    continue
                try:
                    with zf.open(info, "r") as fp:
                        fp.read(1)
                except RuntimeError as e:
                    s = str(e).lower()
                    if "password" in s or "encrypted" in s:
                        return f"加密 ZIP（需密码）: {e!s}"
                except NotImplementedError as e:  # e.g. compression method
                    return f"无法解压: {e!s}"
    except zipfile.BadZipFile as e:
        return f"非法或损坏的 ZIP: {e!s}"
    except Exception as e:
        return f"{type(e).__name__}: {e!s}"
    return None


def _check_excel_unreadable(xlsx_path: str) -> str | None:
    """若 Excel 无法读取（如加密、损坏），返回说明；否则 None。"""
    try:
        pl = xlsx_path.lower()
        if pl.endswith(".xlsx"):
            pd.read_excel(xlsx_path, nrows=0, engine="openpyxl")
        elif pl.endswith(".xls") and not pl.endswith(".xlsx"):
            pd.read_excel(xlsx_path, nrows=0, engine="xlrd")
        else:
            return None
    except Exception as e:
        s = (str(e) or repr(e)).lower()
        if "password" in s or "encrypted" in s or "workbook" in s:
            return f"可能带密码或无法解析: {type(e).__name__}: {e!s}"
        return f"无法打开: {type(e).__name__}: {e!s}"
    return None


def process_all(input_dir, output_file):
    print(f"开始处理目录: {input_dir}")
    all_results: list = []
    failed_items: list[str] = []
    # 1) 仅扫描、提示：加密的 zip（DB 常提供 BILLSTAT_*.zip，需本机先解压到 PDF 再跑）
    zip_files = glob.glob(os.path.join(input_dir, "**", "*.zip"), recursive=True)
    for zpath in zip_files:
        msg = _check_zip_needs_password(zpath)
        if msg:
            line = f"[DB] 警告: {os.path.basename(zpath)} — {msg}（请先解压为 PDF 或将密码写入解压流程。）"
            print(line)
            failed_items.append(line)

    # 1b) 若目录内有 xlsx 但解析逻辑未用，则至少提示打不开的
    for xpath in glob.glob(os.path.join(input_dir, "**", "*.xlsx"), recursive=True) + glob.glob(
        os.path.join(input_dir, "**", "*.xls"), recursive=True
    ):
        if "~$" in os.path.basename(xpath):
            continue
        xmsg = _check_excel_unreadable(xpath)
        if xmsg:
            line = f"[DB] 警告: Excel 未读入明细（可能加密） {os.path.basename(xpath)} — {xmsg}"
            print(line)
            failed_items.append(line)

    pdf_files = glob.glob(os.path.join(input_dir, '**', '*.pdf'), recursive=True)
    for pdf_file in pdf_files:
        try:
            if "BRANCH_SUMMARY" in os.path.basename(pdf_file):
                all_results.extend(parse_summary_pdf(pdf_file))
            else:
                all_results.extend(parse_account_pdf(pdf_file))
        except Exception as e:
            s = str(e) or type(e).__name__
            if "password" in s.lower() or "encrypted" in s.lower() or "decrypt" in s.lower():
                line = f"[DB] 警告: PDF 可能加密/需密码 {os.path.basename(pdf_file)} — {type(e).__name__}: {e!r}"
            else:
                line = f"[DB] 警告: PDF 解析失败 {os.path.basename(pdf_file)} — {type(e).__name__}: {e!r}"
            print(line)
            failed_items.append(line)

    df = pd.DataFrame(all_results)
    if not df.empty:
        columns_order = ['Service Type', 'Account No', 'Product Description', 'Item Count', 'Tariff CCY', 'Amount', 'Avg Per Item', 'Source']
        # 补齐可能缺少的列
        for col in columns_order:
            if col not in df.columns:
                df[col] = None
        df = df[columns_order]
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                
                import openpyxl
                max_col = openpyxl.utils.get_column_letter(worksheet.max_column)
                max_row = worksheet.max_row
                
                worksheet.auto_filter.ref = f"A1:{max_col}{max_row}"
            
            print(f"\n提取完成！共提取 {len(df)} 条明细。")
            print(f"结果已保存至: {output_file}")
        except Exception as e:
            print(f"保存 Excel 时出错: {e}")
    else:
        print("\n未提取到任何明细数据。")
    if failed_items:
        print("\n[DB] 未成功解析/需人工处理的文件清单：")
        for it in failed_items:
            print("  -", it)

if __name__ == "__main__":
    input_directory = r"E:\2月成本分摊模拟\202602\2026.02账单"
    output_excel_path = r"e:\Desktop\demo\DB账单汇总.xlsx"
    process_all(input_directory, output_excel_path)
