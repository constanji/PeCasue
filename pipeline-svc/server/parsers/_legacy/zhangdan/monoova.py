import os
import glob
from datetime import datetime
from pathlib import Path
from typing import Any
import re

'Monoova Tax Invoice 输出列与正则。'

OUTPUT_COLUMNS = ['Reference', 'Date', 'DESCRIPTION', 'VOLUME', 'UNIT PRICE', 'TOTAL CHARGE', 'TOTAL CHARGE CCY']

META_COLUMNS = ['source_pdf', 'source_stem', 'page']

QTY_LINE_RE = re.compile('^\\d[\\d,]*$')

INVOICE_NUM_RE = re.compile('\\b(INV\\d+)\\b', re.IGNORECASE)

DATE_DDMMYYYY_RE = re.compile('Date:\\s*(\\d{1,2}/\\d{1,2}/\\d{4})', re.IGNORECASE)

DEFAULT_CCY = 'AUD'

'从 PDF 文本与文件名解析 Reference、Date。'

def stem_from_path(path: Path) -> str:
    return path.stem

def _ddmmyyyy_to_period_label(s: str) -> str:
    """31/01/2026 -> January 2026（与其它账单 Date 列一致）。"""
    parts = s.strip().split('/')
    if len(parts) != 3:
        return ''
    (d, m, y) = (int(parts[0]), int(parts[1]), int(parts[2]))
    try:
        dt = datetime(y, m, d)
        return dt.strftime('%B %Y')
    except ValueError:
        return ''

def extract_invoice_reference(text: str, stem: str) -> str:
    m = INVOICE_NUM_RE.search(text)
    if m:
        return m.group(1).upper()
    m2 = INVOICE_NUM_RE.search(stem)
    return m2.group(1).upper() if m2 else stem.strip()[:40]

def extract_period_date(text: str, stem: str) -> str:
    m = DATE_DDMMYYYY_RE.search(text)
    if m:
        return _ddmmyyyy_to_period_label(m.group(1))
    m2 = re.search('(January|February|March|April|May|June|July|August|September|October|November|December)\\s*[-–]\\s*(\\d{4})', stem, re.I)
    if m2:
        return f'{m2.group(1).title()} {m2.group(2)}'
    return ''

def strip_aud(s: str) -> str:
    return s.replace('$', '').replace(',', '').strip()

'判断整段文本是否像 Monoova Tax Invoice。'

def looks_like_monoova_invoice(text: str) -> bool:
    t = text.lower()
    if 'monoova' not in t:
        return False
    if 'account-to-account' in t or 'tax invoice' in t:
        return True
    return 'inv' in t and 'quantity' in t and ('rate' in t) and ('amount' in t)

'去重、剔除零金额。'

def parse_total_charge_value(rec: dict[str, Any]) -> float | None:
    raw = str(rec.get('TOTAL CHARGE', '') or '').strip().replace(',', '')
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None

def drop_zero_total_charge_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in rows:
        v = parse_total_charge_value(r)
        if v is not None and v == 0.0:
            continue
        out.append(r)
    return out

def dedupe_by_business_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_file: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_file.setdefault(r.get('source_pdf', ''), []).append(r)
    out: list[dict[str, Any]] = []
    for (_path, group) in by_file.items():
        seen: set[tuple[str, ...]] = set()
        for r in group:
            key = tuple((str(r.get(k, '') or '') for k in OUTPUT_COLUMNS))
            if key in seen:
                continue
            seen.add(key)
            out.append(r)
    return out

'解析 Monoova 明细表（Description 多行 + Quantity + Rate + Amount）。'

def _find_amount_header_index(lines: list[str]) -> int | None:
    for (i, raw) in enumerate(lines):
        if raw.strip() == 'Amount':
            return i
    return None

def _find_subtotal_index(lines: list[str], start: int) -> int:
    for j in range(start, len(lines)):
        s = lines[j].strip()
        if s.startswith('Subtotal') and 'Amount' in s:
            return j
    return len(lines)

def parse_line_items(text: str, stem: str) -> list[dict[str, str]]:
    """
    PyMuPDF 典型行序：若干描述行 → 数量 → 空行/$单价 → $金额。
    """
    lines = text.splitlines()
    hi = _find_amount_header_index(lines)
    if hi is None:
        return []
    start = hi + 1
    end = _find_subtotal_index(lines, start)
    ref = extract_invoice_reference(text, stem)
    period = extract_period_date(text, stem)
    rows: list[dict[str, str]] = []
    desc_buf: list[str] = []
    i = start
    while i < end:
        raw = lines[i]
        line = raw.strip()
        if not line:
            i += 1
            continue
        if QTY_LINE_RE.match(line):
            qty = line.replace(',', '')
            i += 1
            rate_s = ''
            amt_s = ''
            while i < end:
                s = lines[i].strip()
                if s.startswith('$'):
                    if not rate_s:
                        rate_s = strip_aud(s)
                    elif not amt_s:
                        amt_s = strip_aud(s)
                        i += 1
                        break
                i += 1
            desc = ' / '.join(desc_buf) if desc_buf else ''
            desc_buf = []
            rec = {k: '' for k in OUTPUT_COLUMNS}
            rec['Reference'] = ref
            rec['Date'] = period
            rec['DESCRIPTION'] = desc
            rec['VOLUME'] = qty
            rec['UNIT PRICE'] = rate_s
            rec['TOTAL CHARGE'] = amt_s
            rec['TOTAL CHARGE CCY'] = DEFAULT_CCY
            rows.append(rec)
            continue
        desc_buf.append(line)
        i += 1
    return rows

def attach_meta(rows: list[dict[str, str]], source_pdf: str, source_stem: str, page: int) -> list[dict[str, Any]]:
    meta = {'source_pdf': source_pdf, 'source_stem': source_stem, 'page': page}
    return [{**meta, **r} for r in rows]

'从 Monoova Tax Invoice PDF 提取明细行。'

def extract_from_pdf(path: Path) -> list[dict[str, Any]]:
    try:
        import fitz
    except ImportError as e:
        raise RuntimeError('需要安装 pymupdf：pip install pymupdf') from e
    path = path.expanduser().resolve()
    stem = stem_from_path(path)
    out: list[dict[str, Any]] = []
    doc = fitz.open(str(path))
    try:
        for pi in range(len(doc)):
            text = doc[pi].get_text('text') or ''
            if not text.strip():
                continue
            if not looks_like_monoova_invoice(text):
                continue
            rows = parse_line_items(text, stem)
            if not rows:
                continue
            out.extend(attach_meta(rows, str(path), stem, pi + 1))
    finally:
        doc.close()
    return out



import pandas as pd
from typing import Iterable

def iter_pdfs(root: Path, recursive: bool) -> Iterable[Path]:
    if root.is_file() and root.suffix.lower() == ".pdf":
        yield root
        return
    pattern = "**/*.pdf" if recursive else "*.pdf"
    yield from sorted(root.glob(pattern))

def main(input_folder: str, output_excel: str) -> None:
    input_path = Path(input_folder).expanduser().resolve()
    out = Path(output_excel).expanduser().resolve()
    
    pdfs = list(iter_pdfs(input_path, True))
    if not pdfs:
        print(f"未找到 PDF: {input_path}")
        return
        
    print(f"找到 {len(pdfs)} 个 PDF 文件，正在开始解析...")
    records = []
    for p in pdfs:
        try:
            records.extend(extract_from_pdf(p))
        except Exception as e:
            print(f"[跳过] {p}: {e}")
            
    records = dedupe_by_business_columns(records)
    records = drop_zero_total_charge_rows(records)
    
    omit = {"source_pdf", "page"}
    base_cols = META_COLUMNS + OUTPUT_COLUMNS
    out_cols = [c for c in base_cols if c not in omit]
    out_cols.append("source")
    
    if not records:
        print("未能提取到任何有效数据。")
        return
        
    df = pd.DataFrame(records)
    for c in base_cols:
        if c not in df.columns:
            df[c] = ""
    df["source"] = df["source_pdf"].map(lambda p: Path(str(p)).name if p else "")
    df = df[out_cols]
    
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            import openpyxl
            max_col = openpyxl.utils.get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = f"A1:{max_col}{worksheet.max_row}"
        print(f"成功提取了 {len(df)} 条明细记录，并已保存至: {out}")
    except Exception as e:
        print(f"保存 Excel 时出错: {e}")

if __name__ == "__main__":
    main(r"e:\Desktop\demo\monoova_invoice", r"e:\Desktop\demo\monoova_summary.xlsx")
