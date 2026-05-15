import os
import glob
import re
import pandas as pd
import pdfplumber

"""
    xendit银行账单解析工具
"""

def parse_xendit_pdf(pdf_path):
    results = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                    
                # 寻找表头
                lines = text.split('\n')
                parsing = False
                
                # 提取基本信息
                account_npwp = ""
                issuing_date = ""
                invoice_period = ""
                invoice_no = ""
                
                for line in lines:
                    if "NPWP" in line:
                        match = re.search(r'NPWP\s*:\s*([\w\d]+)', line)
                        if match:
                            account_npwp = match.group(1).strip()
                    if "Issuing Date" in line:
                        match = re.search(r'Issuing Date\s*:\s*(.+)', line)
                        if match:
                            issuing_date = match.group(1).strip()
                    if "Invoice Period" in line:
                        match = re.search(r'Invoice Period\s*:\s*(.+)', line)
                        if match:
                            invoice_period = match.group(1).strip()
                    if "Invoice No" in line:
                        match = re.search(r'Invoice No\s*:\s*(.+)', line)
                        if match:
                            invoice_no = match.group(1).strip()
                
                for line in lines:
                    if "CATEGORY PRODUCT UNIT PRICE" in line:
                        parsing = True
                        continue
                    
                    if parsing:
                        if line.startswith("Total Amount Payable") or line.startswith("Tax Base"):
                            parsing = False
                            break
                            
                        # 处理数据行
                        # 示例1: Money In Virtual Account BNI Fixed Fees IDR 4,000 16711 IDR 1,011,060,095,116 IDR 66,844,000
                        # 示例2: Virtual Account BRI Fixed Fees IDR 4,000 1 IDR 2,000 IDR 4,000
                        # 注意：Category (Account) 只有在第一行才出现 (如 Money In, Money Out)
                        
                        parts = line.split()
                        if len(parts) < 6:
                            continue
                            
                        # 从右往左解析
                        # 最后三个是 Fee Amount (e.g. IDR 66,844,000)
                        # FEE AMOUNT
                        fee_amount_str = parts[-1].replace(',', '')
                        fee_ccy = parts[-2]
                        
                        # TRANSACTION VOLUME (我们不需要提取这个金额，但要跳过它)
                        # e.g. IDR 1,011,060,095,116
                        # 这个可能是变长的，因为有时候没有 Transaction Volume。
                        # 我们通过寻找 NO. OF TRANSACTIONS (Volume) 也就是孤立的数字来定位
                        
                        # 寻找从右到左的第一个孤立数字，即为 NO. OF TRANSACTIONS
                        volume_idx = -1
                        for i in range(len(parts)-3, -1, -1):
                            # 如果遇到了 "IDR" 或 "USD" 等货币单位，说明已经越过了 volume 进入到 transaction volume 的金额部分了，需要跳过它
                            if parts[i] in ["IDR", "USD", "EUR", "SGD", "PHP"]:
                                continue
                            # 只有纯数字才可能是 volume
                            if re.match(r'^[\d,]+$', parts[i]):
                                # 如果这个数字前面是货币单位，那它可能是金额而不是 volume
                                if i > 0 and parts[i-1] in ["IDR", "USD", "EUR", "SGD", "PHP"]:
                                    continue
                                volume_idx = i
                                break
                                
                        if volume_idx == -1:
                            continue
                            
                        volume = int(parts[volume_idx].replace(',', ''))
                        
                        # UNIT PRICE (前两个元素是货币单位和单价 e.g. IDR 4,000)
                        unit_price_str = parts[volume_idx - 1].replace(',', '')
                        unit_price_ccy = parts[volume_idx - 2]
                        
                        try:
                            unit_price = float(unit_price_str)
                        except ValueError:
                            # 处理异常情况
                            continue
                            
                        # Product & Category info ends 2 positions before volume (before Unit Price CCY)
                        # The columns are: [Category] Product UnitPriceCCY UnitPrice Volume TransactionVolume FeeAmount
                        # So everything before UnitPriceCCY (volume_idx - 2) is Category + Product
                        prefix_parts = parts[:volume_idx - 2]
                        
                        # Pricing Method: in Xendit, the last two words of the Product info usually indicate the method (e.g., "Fixed Fees")
                        pricing_method = " ".join(prefix_parts[-2:]) if len(prefix_parts) >= 2 else ""
                        
                        # 检查前缀中是否包含已知 Category，例如 Money In, Money Out
                        if len(prefix_parts) >= 2 and " ".join(prefix_parts[:2]) in ["Money In", "Money Out"]:
                            description = " ".join(prefix_parts[2:])
                        else:
                            description = " ".join(prefix_parts)
                            
                        results.append({
                            "Account": account_npwp,
                            "Issuing Date": issuing_date,
                            "Invoice Period": invoice_period,
                            "Invoice No": invoice_no,
                            "Description": description,
                            "Pricing Method": pricing_method,
                            "Volume": volume,
                            "Unit\nPrice": unit_price,
                            "Unit Price CCY": unit_price_ccy,
                            "Charge in Invoice CCY": float(fee_amount_str),
                            "Invoice CCY": fee_ccy,
                            "Taxable": "",  # 模板里有这个字段，但PDF未明确指出，置空
                            "来源文件": os.path.basename(pdf_path)
                        })
    except Exception as e:
        print(f"解析文件 {pdf_path} 失败: {e}")
        
    return results

def process_all(input_dir, output_file):
    print(f"开始处理目录: {input_dir}")
    all_results = []
    pdf_files = glob.glob(os.path.join(input_dir, '**', '*.pdf'), recursive=True)
    
    for pdf_file in pdf_files:
        res = parse_xendit_pdf(pdf_file)
        if res:
            all_results.extend(res)
            
    df = pd.DataFrame(all_results)
    
    if not df.empty:
        columns_order = ['Account', 'Issuing Date', 'Invoice Period', 'Invoice No', 'Description', 'Pricing Method', 'Volume', 'Unit\nPrice', 
                         'Unit Price CCY', 'Charge in Invoice CCY', 'Invoice CCY', 'Taxable', '来源文件']
                         
        for col in columns_order:
            if col not in df.columns:
                df[col] = None
        df = df[columns_order]
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                
                import openpyxl
                max_col = openpyxl.utils.get_column_letter(worksheet.max_column)
                max_row = worksheet.max_row
                
                worksheet.auto_filter.ref = f"A1:{max_col}{max_row}"
                
            print(f"\n提取完成！共处理了 {len(pdf_files)} 个文件，提取了 {len(df)} 条明细。")
            print(f"结果已保存至: {output_file}")
        except Exception as e:
            print(f"保存 Excel 时出错: {e}")
    else:
        print("\n未提取到任何明细数据。")

if __name__ == "__main__":
    input_directory = r"e:\Desktop\demo\xendit账单"
    output_excel_path = r"e:\Desktop\demo\2026.2xendit.xlsx"
    process_all(input_directory, output_excel_path)
